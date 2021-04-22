import pandas as pd
import numpy as np
import mf_lib as mflib
import os
from math import ceil
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import quote_sheetname, get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from shutil import copyfile
import json


with open(r'dependencies\services\dhl_service_hash.json', 'r') as f:
    service = json.load(f)
print(service)
file = 'shopify shipped orders637473383315144984.xlsx'.lower()
# file = 'shipstation 28eb2807-fe05-4cff-b247-1dd6c55c37dd.csv'.lower()
path = rf'Manifests\{file}'
sw = file.split()[0]
format_hash = mflib.ManifestFormat.format_hash
orderno = format_hash['order_no'][sw]  # not required
shipdate = format_hash['date'][sw]
weight = format_hash['weight'][sw]
sv = format_hash['service'][sw]  # not required
address = format_hash['address'][sw]
address1 = format_hash['address1'].get(sw)
current_price = format_hash['current_price'].get(sw)  # optional
insured_parcel = format_hash['insured_parcel'].get(sw)  # optional
dim1 = format_hash['dim1'].get(sw)  # optional
dim2 = format_hash['dim2'].get(sw)  # optional
dim3 = format_hash['dim3'].get(sw)  # optional
type = {'str': str, 'float': float, 'int': pd.Int64Dtype(), 'bool': bool}
# columns = [shipdate['header'], weight['header'], sv['header'], country['header'], zip['header']]
# dtype = [shipdate['format'], weight['format'], sv['format'], country['format'], zip['format']]
# headers = ['shipdate', 'weight', 'service', 'country', 'zip']


def create_df(path, columns, dtype, headers):
    print(path, columns, dtype, headers)
    f_ext = path.rsplit('.', 1)[1]
    if f_ext == 'xlsx':
        df = pd.read_excel(path, usecols=columns, dtype=dtype)
    elif f_ext == 'csv':
        df = pd.read_csv(path, usecols=columns, dtype=dtype)
    print(df.head(3))
    # parse_dates=[columns[1]], date_parser=dateparse
    df = df[columns]
    df = df.rename(columns=headers)
    df.dropna(how='all', inplace=True)
    df.dropna(subset=['weight'], inplace=True)
    print(df.head(5))
    df['shipdate'] = pd.to_datetime(df['shipdate'])
    pd.set_option('display.max_columns', None)
    return df


def sellercloud_shipbridge(path, col_set, df):
    print(col_set)
    columns = []
    dtype = {}
    headers = {}
    not_found = []
    weight_name = None
    # [orderno['header'], shipdate['header'], weight['header'],
    # sv['header'], address['header'], current_price['header']]
    pot_columns = {'orderno': orderno, 'shipdate': shipdate, 'weight': weight,
                   'service': sv, 'address': address, 'price': current_price}
    for pot_c in pot_columns.keys():
        for col in pot_columns[pot_c]['header']:
            print(col)
            print(pot_c)
            found = False
            if col in col_set:
                if pot_c != 'weight':
                    dtype[col] = pot_columns[pot_c]['format']
                else:
                    weight_name = col
                headers[col] = pot_c
                found = True
                break
        if not found:
            not_found.append(pot_c)
    # optional_pot_columns = {'dim1': dim1, 'dim2': dim2, 'dim3': dim3}
    # for pot_c in optional_pot_columns.keys():
    #     for col in optional_pot_columns[pot_c]['header']:
    #         found = False
    #         if col in col_set:
    #             dtype[col] = optional_pot_columns[pot_c]['format']
    #             headers[col] = pot_c
    #             found = True
    #             break
    columns = list(headers.keys())
    print(columns)
    print(dtype)
    print(headers)
    # sv_main, sv_alt0, sv_alt1 = False, False, False
    # if 'service' in headers.values():
    #     sv_main = True
    # else:
    #     if sv['header_alt'][0][0] in col_set:
    #         sv_alt0 = True
    #         columns.append(sv['header_alt'][0][0])
    #         dtype[columns[-1]] = type[sv['format']]
    #         headers[columns[-1]] = 'service_provider'
    #     elif sv['header_alt'][0][1] in col_set:
    #         sv_alt0 = True
    #         columns.append(sv['header_alt'][0][1])
    #         dtype[columns[-1]] = type[sv['format']]
    #         headers[columns[-1]] = 'service_provider'
    #     if sv['header_alt'][1][0] in col_set:
    #         sv_alt1 = True
    #         columns.append(sv['header_alt'][1][0])
    #         dtype[columns[-1]] = type[sv['format']]
    #         headers[columns[-1]] = 'service_code'
    #     elif sv['header_alt'][1][1] in col_set:
    #         sv_alt1 = True
    #         columns.append(sv['header_alt'][1][1])
    #         dtype[columns[-1]] = type[sv['format']]
    #         headers[columns[-1]] = 'service_code'
    #     if not (sv_alt0 and sv_alt1):
    #         return  # If service name not included, we can't process the file. Service provider name is not essential
    if weight_name:
        weight_test = str(df[weight_name].iloc[0])
        if weight_test.replace('.', '').isnumeric():
            dtype[weight_name] = weight['format']
            print(weight_test)
            df = create_df(path, columns, dtype, headers)
        elif 'oz' in weight_test or 'lb' in weight_test or 'lbs' in weight_test:
            print(weight_test)
            dtype[weight_name] = 'str'
            df = create_df(path, columns, dtype, headers)
            df['weight'] = df.apply(lambda row: mflib.w_lbs_or_w_oz(row['weight']), axis=1)
    else:
        df = create_df(path, columns, dtype, headers)
    # At this point we either have a service column with vendor and service code,
    # or a service code column with an optional service provider column.
    # If service provider is given, concatenate with service code.
    # if sv_alt0:
    #     df['service'] = df[['service_provider', 'service_code']].agg(' '.join, axis=1)
    #     del df['service_provider']
    #     del df['service_code']
    print(df.head(5))
    return df

    columns = [orderno['header'], shipdate['header'], weight['header'],
               sv['header'], address['header'], current_price['header']]
    for col in columns:
        if col not in col_set:
            print(f'{col} not found in file')
            return
    dtype = {columns[0]: type[orderno['format']], columns[1]: type[shipdate['format']], columns[2]: type[weight['format']],
             columns[3]: type[sv['format']], columns[4]: type[address['format']], columns[5]: type[current_price['format']]}
    headers = {columns[0]: 'orderno', columns[1]: 'shipdate', columns[2]: 'weight',
               columns[3]: 'service', columns[4]: 'address', columns[5]: 'price'}
    df = create_df(path, columns, dtype, headers)
    if weight['parse'] == 'w_lbs_or_w_oz':
        df.weight = df.apply(lambda row: mflib.w_lbs_or_w_oz(row['weight']), axis=1)
    df.drop(df[df['weight'] == 0].index, inplace=True)
    df[['zip', 'country']] = df.apply(lambda row: mflib.add_to_zip_ctry(row.address), axis=1, result_type='expand')
    return df


def shipstation(path, col_set, df):
    columns = []
    dtype = {}
    headers = {}
    not_found = []
    weight_name = None
    pot_columns = {'orderno': orderno, 'shipdate': shipdate, 'weight': weight,
                   'service': sv, 'zip': address, 'country': address1, 'price': current_price,
                   'insured': insured_parcel}
    for pot_c in pot_columns.keys():
        for col in pot_columns[pot_c]['header']:
            found = False
            if col in col_set:
                if pot_c != 'weight':
                    dtype[col] = pot_columns[pot_c]['format']
                else:
                    weight_name = col
                headers[col] = pot_c
                found = True
                break
        if not found:
            not_found.append(pot_c)
    optional_pot_columns = {'dim1': dim1, 'dim2': dim2, 'dim3': dim3}
    for pot_c in optional_pot_columns.keys():
        for col in optional_pot_columns[pot_c]['header']:
            found = False
            if col in col_set:
                dtype[col] = optional_pot_columns[pot_c]['format']
                headers[col] = pot_c
                found = True
                break
    columns = list(headers.keys())
    print(columns)
    print(dtype)
    print(headers)
    sv_main, sv_alt0, sv_alt1 = False, False, False
    if 'service' in headers.values():
        sv_main = True
    else:
        if sv['header_alt'][0][0] in col_set:
            sv_alt0 = True
            columns.append(sv['header_alt'][0][0])
            dtype[columns[-1]] = type[sv['format']]
            headers[columns[-1]] = 'service_provider'
        elif sv['header_alt'][0][1] in col_set:
            sv_alt0 = True
            columns.append(sv['header_alt'][0][1])
            dtype[columns[-1]] = type[sv['format']]
            headers[columns[-1]] = 'service_provider'
        if sv['header_alt'][1][0] in col_set:
            sv_alt1 = True
            columns.append(sv['header_alt'][1][0])
            dtype[columns[-1]] = type[sv['format']]
            headers[columns[-1]] = 'service_code'
        elif sv['header_alt'][1][1] in col_set:
            sv_alt1 = True
            columns.append(sv['header_alt'][1][1])
            dtype[columns[-1]] = type[sv['format']]
            headers[columns[-1]] = 'service_code'
        if not (sv_alt0 and sv_alt1):
            return  # If service name not included, we can't process the file. Service provider name is not essential
    if weight_name:
        weight_test = str(df[weight_name].iloc[0])
        if weight_test.replace('.', '').isnumeric():
            dtype[weight_name] = weight['format']
            df = create_df(path, columns, dtype, headers)
        elif 'oz' in weight_test or 'lb' in weight_test or 'lbs' in weight_test:
            dtype[weight_name] = 'str'
            df = create_df(path, columns, dtype, headers)
            df['weight'] = df.apply(lambda row: mflib.w_lbs_or_w_oz(row['weight']), axis=1)
    else:
        df = create_df(path, columns, dtype, headers)
    # At this point we either have a service column with vendor and service code,
    # or a service code column with an optional service provider column.
    # If service provider is given, concatenate with service code.
    if sv_alt0:
        df['service'] = df[['service_provider', 'service_code']].agg(' '.join, axis=1)
        del df['service_provider']
        del df['service_code']
    print(df.head(5))
    return df


def shopify(path, col_set, df):
    columns = []
    dtype = {}
    headers = {}
    not_found = []
    weight_name = None
    pot_columns = {'orderno': orderno, 'shipdate': shipdate, 'weight': weight,
                   'service': sv, 'zip': address, 'country': address1, 'price': current_price}
    for pot_c in pot_columns.keys():
        for col in pot_columns[pot_c]['header']:
            found = False
            if col in col_set:
                if pot_c != 'weight':
                    dtype[col] = pot_columns[pot_c]['format']
                else:
                    weight_name = col
                headers[col] = pot_c
                found = True
                break
        if not found:
            not_found.append(pot_c)
    # optional_pot_columns = {'dim1': dim1, 'dim2': dim2, 'dim3': dim3}
    # for pot_c in optional_pot_columns.keys():
    #     for col in optional_pot_columns[pot_c]['header']:
    #         found = False
    #         if col in col_set:
    #             dtype[col] = optional_pot_columns[pot_c]['format']
    #             headers[col] = pot_c
    #             found = True
    #             break
    columns = list(headers.keys())
    print(columns)
    print(dtype)
    print(headers)
    # dtype = {columns[2]: type[address['format']],
    #          columns[3]: type[address['format']], columns[4]: type[current_price['format']],
    #          columns[5]: type[dimensions['format']], columns[6]: type[dimensions['format']],
    #          columns[7]: type[dimensions['format']], columns[8]: type[insured_parcel['format']]}
    # headers = {columns[0]: 'shipdate', columns[1]: 'weight',
    #            columns[2]: 'zip', columns[3]: 'country', columns[4]: 'price',
    #            columns[5]: 'dim0', columns[6]: 'dim1', columns[7]: 'dim2', columns[8]: 'insured'}
    # if (dimensions['header0'] not in col_set):
    #     del dtype[columns[-4]]
    #     del dtype[columns[-3]]
    #     del dtype[columns[-2]]
    #     del headers[columns[-4]]
    #     del headers[columns[-3]]
    #     del headers[columns[-2]]
    #     print(columns)
    #     columns = columns[:-4]+[columns[-1]]
    #     print(columns)
    if weight_name:
        print(weight_name)
        df.dropna(subset=[weight_name], inplace=True)
        for i, row in df.iterrows():
            weight_test = str(row[weight_name])
            if weight_test.replace('.', '').isnumeric():
                dtype[weight_name] = weight['format']
                df = create_df(path, columns, dtype, headers)
                break
            elif 'oz' in weight_test or 'lb' in weight_test or 'lbs' in weight_test:
                dtype[weight_name] = 'str'
                df = create_df(path, columns, dtype, headers)
                df['weight'] = df.apply(lambda row: mflib.w_lbs_or_w_oz(row['weight']), axis=1)
                break
    else:
        df = create_df(path, columns, dtype, headers)
    # At this point we either have a service column with vendor and service code,
    # or a service code column with an optional service provider column.
    # If service provider is given, concatenate with service code.
    # df['insured'] = (df['insured'] > 0)
    # print(df.head())
    return df


def dateparse(x): return datetime.strptime(x, shipdate['parse'])


# for dt in dtype:
#     print(dt, dtype[dt])
# df = pd.read_excel(path, parse_dates=[columns[0]], date_parser=dateparse, usecols=columns, dtype=dtype)
f_ext = path.rsplit('.', 1)[1]
w_file = fr'Manifests\Manifest {file.rsplit(".", 1)[0]} s.xlsx'
if not os.path.exists(w_file):
    if f_ext == 'xlsx':
        df = pd.read_excel(path, nrows=5)
    elif f_ext == 'csv':
        df = pd.read_csv(path, nrows=5)
    col_set = set()
    for i in df.columns:
        col_set.add(i)
    df = locals()[sw](path, col_set, df)
    df[['country', 'zone', 'weight threshold', 'sugg. service', 'bill weight', '2021 tier 1', '2021 tier 2', '2021 tier 3', '2021 tier 4',
        '2021 tier 5', '2021 DHL', '2021 USPS', 'shipdate DHL', 'shipdate USPS']] = df.apply(
        lambda row: mflib.row_to_rate(row), axis=1, result_type='expand')
    df.weight = df.apply(lambda row: ceil(row.weight) if row.weight < 16 else ceil(row.weight/16)*16, axis=1)
    df.loc[~df['bill weight'].isna(), 'weight'] = df['bill weight']
    del df['bill weight']
    writer = pd.ExcelWriter(w_file, engine='openpyxl')
    df.to_excel(writer, sheet_name='Sheet1', freeze_panes=(1, 0), index=False)
    df_zone_not_found = df[df['zone'].isna()][['orderno', 'zip', 'country']]
    subset = ['service', 'weight threshold', 'country']
    df.loc[df['zone'].str.len() != 7, 'country'] = 'Intl'
    df_unique_services = df[['service', 'weight threshold', 'country',
                             'sugg. service']].drop_duplicates(subset=subset, inplace=False)
    df_unique_services['Actual Service'] = np.nan
    df_unique_services.sort_values(by=['country', 'weight threshold', 'service'],
                                   ascending=[False, True, True], inplace=True)
    service_list = [[v[0], v[-1]] for v in service.values()]
    service_names = [v[-1] for v in service.values()]
    df_existing_services = pd.DataFrame(service_list, columns=['Service #', 'Service Name'])
    print(df_existing_services.head())
    df_zone_not_found.to_excel(writer, sheet_name='Zone Not Found', freeze_panes=(1, 0), index=False)
    col = 0
    spaces = 2
    for dataframe in [df_unique_services, df_existing_services]:
        dataframe.to_excel(writer, sheet_name='Unique Service Params', startcol=col, index=False)
        col += len(dataframe.columns) + spaces + 1
    writer.save()
    writer.close()
    wb = load_workbook(w_file)
    ws = wb.active

    MIN_WIDTH = 15
    for i, column_cells in enumerate(ws.columns, start=1):
        width = MIN_WIDTH
        ws.column_dimensions[get_column_letter(i)].width = width

    wb.active = 2
    ws = wb.active

    expression = "{0}!$J$2:$J$"+str(len(df_existing_services.index)+1)
    formula = expression.format(quote_sheetname('Unique Service Params'))
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)

    # Optionally set a custom error message
    dv.error = 'Your entry is not in the list'
    dv.errorTitle = 'Invalid Entry'

    # Optionally set a custom prompt message
    dv.prompt = 'Please select service'
    dv.promptTitle = 'Drop Down List'

    # Add the data-validation object to the worksheet
    ws.add_data_validation(dv)

    for i in range(len(df_unique_services.index)):
        dv.add(ws[f'E{i+2}'])
    wb.save(w_file)

    for i, column_cells in enumerate(ws.columns, start=1):
        width = (
            length
            if (length := max(len(str(cell_value) if (cell_value := cell.value) is not None else "")
                              for cell in column_cells)) >= MIN_WIDTH
            else MIN_WIDTH
        ) if i != 5 else len(max(service_names, key=len))
        ws.column_dimensions[get_column_letter(i)].width = width
    wb.save(w_file)

else:
    # df = pd.read_excel(w_file, sheet_name='Sheet1', usecols=list(
    # print(df.head(8))
    df_unique_services = pd.read_excel(w_file, sheet_name='Unique Service Params', usecols=list(
        mflib.cols_service.keys()), dtype=mflib.cols_service)
    df_unique_services.apply(lambda row: mflib.update_services(row), axis=1)

    df = pd.read_excel(w_file, sheet_name='Sheet1', usecols=list(
        mflib.cols_read.keys()), dtype=mflib.cols_read)
    print(df.head(5))
    print(df_unique_services.head(10))
    df[['sugg. service', '2021 tier 1', '2021 tier 2', '2021 tier 3', '2021 tier 4', '2021 tier 5', '2021 DHL', '2021 USPS',
        'shipdate DHL', 'shipdate USPS']] = df.apply(lambda row: mflib.cost_correction(row), axis=1, result_type='expand')
    print(df.head(5))
    # writer = pd.ExcelWriter('testttt.xlsx', engine='openpyxl')
    with pd.ExcelWriter(w_file, engine='openpyxl', mode='a') as writer:
        df.to_excel(writer, sheet_name='Sheet1', index=False)
